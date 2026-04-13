import streamlit as st
import pandas as pd
import io
import re
import uuid
from datetime import datetime, date, time, timedelta
from dateutil import parser as dtparser
import pytz
from openpyxl import load_workbook
from icalendar import Calendar, Event, Timezone, TimezoneStandard, TimezoneDaylight
from streamlit_calendar import calendar as st_calendar

# ==============================================================================
# 1. CONFIGURATION & DESIGN "CESI x APPLE"
# ==============================================================================
st.set_page_config(page_title="Modern Opus", page_icon="🎓", layout="wide")

# Palette CESI (Nouveau Jaune) & Apple Style
CESI_YELLOW = "#f7e34f" 
CESI_BLACK = "#000000"
APPLE_GRAY = "#F5F5F7"
WHITE = "#FFFFFF"

st.markdown(f"""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
    
    /* Global Reset */
    html, body, [class*="css"] {{
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
        background-color: {APPLE_GRAY};
        color: {CESI_BLACK};
    }}
    
    /* STYLE CESI : Surlignage Jaune (Tags) */
    .cesi-tag {{
        background-color: {CESI_YELLOW};
        color: {CESI_BLACK};
        font-weight: 800;
        padding: 2px 8px;
        text-transform: uppercase;
        font-size: 0.9rem;
        display: inline-block;
        margin-right: 10px;
        letter-spacing: 0.5px;
    }}

    /* STYLE CESI : Soulignement épais (Titres) */
    .cesi-title {{
        font-weight: 800;
        font-size: 2.5rem;
        color: {CESI_BLACK};
        position: relative;
        display: inline-block;
        margin-bottom: 10px;
    }}
    .cesi-title::after {{
        content: '';
        display: block;
        width: 60px;
        height: 6px;
        background-color: {CESI_YELLOW};
        margin-top: 5px;
    }}
    
    /* Header Minimaliste */
    .main-header {{
        background-color: {WHITE};
        padding: 2rem 0;
        border-bottom: 1px solid #E5E5E5;
        margin-bottom: 2rem;
        text-align: center;
    }}
    
    /* Cards Metrics (Apple Style) */
    .metric-card {{
        background-color: {WHITE};
        border-radius: 12px;
        padding: 1.5rem;
        box-shadow: 0 4px 12px rgba(0,0,0,0.03);
        text-align: center;
        border: 1px solid #EAEAEA;
        transition: transform 0.2s;
    }}
    .metric-card:hover {{
        transform: translateY(-2px);
    }}
    .metric-value {{
        font-size: 2rem;
        font-weight: 800;
        color: {CESI_BLACK};
    }}
    .metric-label {{
        font-size: 0.85rem;
        font-weight: 600;
        color: #86868b;
        text-transform: uppercase;
        margin-top: 5px;
    }}
    
    /* Tabs Élégants */
    .stTabs [data-baseweb="tab-list"] {{
        gap: 8px;
        background-color: transparent;
        padding-bottom: 10px;
    }}
    .stTabs [data-baseweb="tab"] {{
        height: 45px;
        background-color: {WHITE};
        border-radius: 6px;
        color: #1d1d1f;
        font-weight: 600;
        border: 1px solid #E5E5E5;
        padding: 0 20px;
    }}
    .stTabs [aria-selected="true"] {{
        background-color: {CESI_BLACK};
        color: {WHITE};
        border-color: {CESI_BLACK};
    }}

    /* Buttons */
    .stButton button {{
        background-color: {CESI_BLACK};
        color: {WHITE};
        border-radius: 6px;
        font-weight: 600;
        padding: 0.6rem 1.2rem;
        border: none;
        transition: all 0.2s ease;
    }}
    .stButton button:hover {{
        background-color: {CESI_YELLOW};
        color: {CESI_BLACK};
        transform: scale(1.02);
    }}
    
    /* Dataframes clean */
    [data-testid="stDataFrame"] {{
        border: 1px solid #E5E5E5;
        border-radius: 8px;
        overflow: hidden;
        background-color: white;
    }}
</style>
""", unsafe_allow_html=True)

# ==============================================================================
# 2. LOGIQUE MÉTIER (ROBUSTE & SANS RÉGRESSION)
# ==============================================================================

def normalize_group_label(x):
    if x is None: return None
    try:
        if pd.isna(x): return None
    except: pass
    s = str(x).strip()
    if not s: return None
    m = re.search(r'G\s*\.?\s*(\d+)', s, re.I)
    if m: return f'G {m.group(1)}'
    m2 = re.search(r'^(?:groupe)?\s*(\d+)$', s, re.I)
    if m2: return f'G {m2.group(1)}'
    return s

def is_time_like(x):
    if x is None: return False
    if isinstance(x, (pd.Timestamp, datetime, time)): return True
    s = str(x).strip()
    if not s: return False
    if re.match(r'^\d{1,2}[:hH]\d{2}(\s*[AaPp][Mm]\.?)?$', s): return True
    return False

def to_time(x):
    if x is None: return None
    if isinstance(x, time): return x
    if isinstance(x, pd.Timestamp): return x.to_pydatetime().time()
    if isinstance(x, datetime): return x.time()
    s = str(x).strip()
    if not s: return None
    s2 = s.replace('h', ':').replace('H', ':')
    try:
        dt = dtparser.parse(s2, dayfirst=True)
        return dt.time()
    except: return None

def to_date(x):
    if x is None: return None
    if isinstance(x, pd.Timestamp): return x.to_pydatetime().date()
    if isinstance(x, datetime): return x.date()
    if isinstance(x, date): return x
    s = str(x).strip()
    if not s: return None
    try:
        dt = dtparser.parse(s, dayfirst=True, fuzzy=True)
        return dt.date()
    except: return None

def get_merged_map_from_bytes(xls_bytes, sheet_name):
    wb = load_workbook(io.BytesIO(xls_bytes), data_only=True)
    ws = wb[sheet_name]
    merged_map = {}
    for merged in ws.merged_cells.ranges:
        r1, r2, c1, c2 = merged.min_row, merged.max_row, merged.min_col, merged.max_col
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                merged_map[(r - 1, c - 1)] = (r1 - 1, c1 - 1, r2 - 1, c2 - 1)
    return merged_map

@st.cache_data(show_spinner=False)
def parse_excel_engine(file_content, sheet_names_to_scan):
    """ Moteur de parsing unifié """
    results = {}
    
    for sheet in sheet_names_to_scan:
        try:
            # Renommage P1/P2 pour l'affichage
            promo_name = "P1" if "P1" in sheet.upper() else ("P2" if "P2" in sheet.upper() else sheet)

            df = pd.read_excel(io.BytesIO(file_content), sheet_name=sheet, header=None, engine='openpyxl')
            merged_map = get_merged_map_from_bytes(file_content, sheet)
            
            nrows, ncols = df.shape
            s_rows = [i for i in range(len(df)) if isinstance(df.iat[i,0], str) and re.match(r'^\s*S\s*\d+', df.iat[i,0].strip(), re.I)]
            h_rows = [i for i in range(len(df)) if isinstance(df.iat[i,0], str) and re.match(r'^\s*H\s*\d+', df.iat[i,0].strip(), re.I)]
            
            raw_events = []
            
            for r in h_rows:
                p_candidates = [s for s in s_rows if s <= r]
                if not p_candidates: continue
                p = max(p_candidates)
                date_row, group_row = p + 1, p + 2
                
                date_cols = [c for c in range(ncols) if date_row < nrows and to_date(df.iat[date_row, c]) is not None]
                
                for c in date_cols:
                    for col in (c, c + 1):
                        if col >= ncols: continue
                        
                        try: summary = df.iat[r, col]
                        except: summary = None
                        if pd.isna(summary) or summary is None: continue
                        summary_str = str(summary).strip()
                        if not summary_str: continue
                        
                        # Teachers
                        teachers = []
                        if (r+2) < nrows:
                            for off in range(2, 6):
                                if (r+off) >= nrows: break
                                try: t = df.iat[r+off, col]
                                except: t = None
                                if t and not pd.isna(t) and not is_time_like(t) and not isinstance(t, (int, float)):
                                    s_t = str(t).strip()
                                    if s_t: teachers.append(s_t)
                        teachers = list(dict.fromkeys(teachers))
                        
                        stop_idx = None
                        for off in range(1, 12):
                            idx = r + off
                            if idx >= nrows: break
                            try:
                                if is_time_like(df.iat[idx, col]):
                                    stop_idx = idx
                                    break
                            except: continue
                        if stop_idx is None: stop_idx = min(r+7, nrows)
                        
                        desc_parts = []
                        for idx in range(r+1, stop_idx):
                            if idx >= nrows: break
                            try: cell = df.iat[idx, col]
                            except: cell = None
                            if pd.isna(cell) or cell is None: continue
                            s_cell = str(cell).strip()
                            if not s_cell or to_date(cell) is not None: continue
                            if s_cell in teachers or s_cell == summary_str: continue
                            desc_parts.append(s_cell)
                        desc_text = " | ".join(dict.fromkeys(desc_parts))
                        
                        start_val, end_val = None, None
                        for off in range(1, 13):
                            idx = r + off
                            if idx >= nrows: break
                            try: v = df.iat[idx, col]
                            except: v = None
                            if is_time_like(v):
                                if start_val is None: start_val = v
                                elif end_val is None and v != start_val:
                                    end_val = v; break
                        if start_val is None or end_val is None: continue
                        start_t, end_t = to_time(start_val), to_time(end_val)
                        if start_t is None or end_t is None: continue
                        
                        d = to_date(df.iat[date_row, c])
                        if d is None: continue
                        dtstart, dtend = datetime.combine(d, start_t), datetime.combine(d, end_t)
                        
                        gl = normalize_group_label(df.iat[group_row, col] if group_row < nrows else None)
                        gl_next = normalize_group_label(df.iat[group_row, col+1] if (col+1) < ncols and group_row < nrows else None)
                        
                        groups = set()
                        is_left = (col == c)
                        if is_left:
                            merged_here = merged_map.get((r, col))
                            merged_right = merged_map.get((r, col+1))
                            if merged_here and merged_right and merged_here == merged_right:
                                if gl: groups.add(gl)
                                if gl_next: groups.add(gl_next)
                            else:
                                if gl: groups.add(gl)
                        else:
                            if gl: groups.add(gl)
                            
                        raw_events.append({
                            'summary': summary_str,
                            'teachers': set(teachers),
                            'descriptions': set([desc_text]) if desc_text else set(),
                            'start': dtstart,
                            'end': dtend,
                            'groups': groups
                        })
            
            merged = {}
            for e in raw_events:
                key = (e['summary'], e['start'], e['end'])
                if key not in merged:
                    merged[key] = {
                        'summary': e['summary'],
                        'teachers': set(),
                        'descriptions': set(),
                        'start': e['start'],
                        'end': e['end'],
                        'groups': set()
                    }
                merged[key]['teachers'].update(e.get('teachers', set()))
                merged[key]['descriptions'].update(e.get('descriptions', set()))
                merged[key]['groups'].update(e.get('groups', set()))
            
            final_events = []
            for v in merged.values():
                final_events.append({
                    'summary': v['summary'],
                    'teachers': sorted([t for t in v['teachers'] if t and str(t).lower() not in ['nan','none']]),
                    'description': " | ".join(sorted([d for d in v['descriptions'] if d])),
                    'start': v['start'],
                    'end': v['end'],
                    'groups': sorted(list(v['groups']))
                })
            
            results[promo_name] = final_events
            
        except Exception as e:
            results[sheet] = []
            
    return results

def generate_ics(events, tzname='Europe/Paris', include_prefix=False):
    """ 
    Génère l'ICS.
    Si include_prefix=True (pour exports enseignants), ajoute [P1 G 1] au titre.
    """
    cal = Calendar()
    cal.add('prodid', '-//Modern Opus//FR')
    cal.add('version', '2.0')
    
    tz = Timezone()
    tz.add('TZID', tzname)
    std = TimezoneStandard()
    std.add('DTSTART', datetime(1970, 10, 25, 3, 0, 0))
    std.add('TZOFFSETFROM', timedelta(hours=2))
    std.add('TZOFFSETTO', timedelta(hours=1))
    std.add('RRULE', {'freq': 'yearly', 'bymonth': 10, 'byday': '-1su'})
    tz.add_component(std)
    dst = TimezoneDaylight()
    dst.add('DTSTART', datetime(1970, 3, 29, 2, 0, 0))
    dst.add('TZOFFSETFROM', timedelta(hours=1))
    dst.add('TZOFFSETTO', timedelta(hours=2))
    dst.add('RRULE', {'freq': 'yearly', 'bymonth': 3, 'byday': '-1su'})
    tz.add_component(dst)
    cal.add_component(tz)

    timezone = pytz.timezone(tzname)

    for ev in events:
        e = Event()
        
        # Gestion du titre [P1 G 1] Matière pour l'export enseignant
        summary_text = ev['summary']
        if include_prefix:
            parts = []
            # Ajout Promo
            if ev.get('promo_label'):
                parts.append(ev['promo_label'])
            # Ajout Groupes
            if ev.get('groups'):
                parts.extend(ev['groups'])
            
            if parts:
                prefix = f"[{' '.join(parts)}] "
                summary_text = prefix + summary_text

        e.add('summary', summary_text)
        
        start_dt = ev['start']
        end_dt = ev['end']
        if start_dt.tzinfo is None: start_dt = timezone.localize(start_dt)
        if end_dt.tzinfo is None: end_dt = timezone.localize(end_dt)
        
        e.add('dtstart', start_dt)
        e.add('dtend', end_dt)
        e.add('dtstamp', datetime.now(timezone))
        e.add('uid', str(uuid.uuid4()))
        
        desc_lines = []
        if ev['description']: desc_lines.append(ev['description'])
        if ev['teachers']: desc_lines.append('Enseignant(s): ' + ', '.join(ev['teachers']))
        if ev['groups']: desc_lines.append('Groupes: ' + ', '.join(ev['groups']))
        e.add('description', '\n'.join(desc_lines))
        cal.add_component(e)
        
    return cal.to_ical()

# ==============================================================================
# 3. INTERFACE
# ==============================================================================

st.markdown('<div class="main-header"><div class="cesi-title">Modern Opus</div><br><span style="color:#666; font-weight:500;">Planification Intelligente</span></div>', unsafe_allow_html=True)

uploaded_file = st.file_uploader("Déposer le fichier Excel", type=['xlsx'])

if uploaded_file:
    file_bytes = uploaded_file.read()
    
    try:
        xls = pd.ExcelFile(io.BytesIO(file_bytes), engine='openpyxl')
        all_sheets = xls.sheet_names
        # Filtre strict : EDT + P1/P2
        promo_sheets = [s for s in all_sheets if "EDT" in s.upper() and ("P1" in s.upper() or "P2" in s.upper())]
        maquette_sheet = next((s for s in all_sheets if "maquette" in s.lower()), None)
    except Exception as e:
        st.error(f"Erreur fichier: {e}")
        st.stop()

    if not promo_sheets:
        st.error("Aucune feuille 'EDT P1' ou 'EDT P2' détectée.")
        st.stop()

    with st.spinner('Chargement...'):
        events_map = parse_excel_engine(file_bytes, promo_sheets)

    # Agrégation
    all_events_flat = []
    teachers_set = set()
    subjects_set = set()
    
    for promo_name, evs in events_map.items():
        for e in evs:
            e['promo_label'] = promo_name
            all_events_flat.append(e)
            for t in e['teachers']: teachers_set.add(t)
            subjects_set.add(e['summary'])
    
    # --- Metrics ---
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(f'<div class="metric-card"><div class="metric-value">{len(all_events_flat)}</div><div class="metric-label">Séances</div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="metric-card"><div class="metric-value">{len(promo_sheets)}</div><div class="metric-label">Promos</div></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="metric-card"><div class="metric-value">{len(teachers_set)}</div><div class="metric-label">Enseignants</div></div>', unsafe_allow_html=True)
    
    st.write("") 

    tab_cal, tab_mail, tab_stats, tab_exam, tab_export, tab_maquette, tab_volume = st.tabs([
        "🗓️ Calendrier", 
        "✉️ Mails",
        "📊 Récapitulatifs",
        "🎓 Examens",
        "📥 Exports", 
        "📐 Maquette",
        "📈 Volume Horaire"
    ])

    # --- TAB 1: CALENDRIER ---
    with tab_cal:
        col_sel, col_view = st.columns([1, 4])
        with col_sel:
            cal_promo = st.selectbox("Promo", list(events_map.keys()), key="cal_p")
            view_mode = st.radio("Vue", ["Semaine", "Mois"])
        
        with col_view:
            cal_events = []
            for ev in events_map.get(cal_promo, []):
                cal_events.append({
                    "title": ev['summary'],
                    "start": ev['start'].isoformat(),
                    "end": ev['end'].isoformat(),
                    "backgroundColor": CESI_YELLOW,
                    "borderColor": CESI_YELLOW,
                    "textColor": CESI_BLACK,
                    "extendedProps": {"description": f"{', '.join(ev['teachers'])}"}
                })
            
            calendar_options = {
                "headerToolbar": {
                    "left": "today prev,next",
                    "center": "title",
                    "right": ""
                },
                "initialView": "timeGridWeek" if view_mode == "Semaine" else "dayGridMonth",
                "slotMinTime": "07:30:00",
                "slotMaxTime": "19:30:00",
                "contentHeight": "auto",
                "locale": "fr",
                "allDaySlot": False,
                "nowIndicator": True,
                "eventBorderColor": "transparent"
            }
            if cal_events:
                st_calendar(events=cal_events, options=calendar_options)

    # --- TAB 2: MAILS ---
    with tab_mail:
        c_m1, c_m2 = st.columns([1, 3])
        sorted_teachers = sorted(list(teachers_set))
        
        with c_m1:
            st.markdown('<div class="cesi-tag">Configuration</div>', unsafe_allow_html=True)
            st.write("")
            if sorted_teachers:
                chosen_teacher = st.selectbox("Destinataire", sorted_teachers)
                politesse = st.radio("Ton", ["Tutoiement", "Vouvoiement"])
            else:
                chosen_teacher = None
                st.info("Pas d'enseignants.")
        
        with c_m2:
            if chosen_teacher:
                t_evs = [e for e in all_events_flat if chosen_teacher in e['teachers']]
                clean_name = chosen_teacher.replace(',', '')
                parts = clean_name.split()
                
                if politesse == "Tutoiement":
                    prenom = parts[1] if len(parts) > 1 else parts[0]
                    intro = f"Bonjour {prenom},\n\nVoici le récapitulatif de tes interventions :"
                    closing = "Bien à toi,"
                else:
                    nom = parts[0]
                    intro = f"Bonjour M./Mme {nom},\n\nVeuillez trouver ci-dessous le récapitulatif de vos interventions :"
                    closing = "Cordialement,"
                
                body = ""
                months = {1:'janvier', 2:'février', 3:'mars', 4:'avril', 5:'mai', 6:'juin', 7:'juillet', 8:'août', 9:'septembre', 10:'octobre', 11:'novembre', 12:'décembre'}
                days = {0:'Lundi', 1:'Mardi', 2:'Mercredi', 3:'Jeudi', 4:'Vendredi', 5:'Samedi', 6:'Dimanche'}

                promos_present = sorted(list(set(e['promo_label'] for e in t_evs)))
                total_h_global = 0
                
                for promo in promos_present:
                    p_evs = [e for e in t_evs if e['promo_label'] == promo]
                    if not p_evs: continue
                    
                    # Pas d'étoiles, format texte brut
                    body += f"\nPour la promo {promo} :\n"
                    by_subj = {}
                    for e in p_evs:
                        by_subj.setdefault(e['summary'], []).append(e)
                    
                    for subj, ev_list in by_subj.items():
                        # Fusion des séances consécutives (Même matière, même journée, fin == début)
                        # 1. Tri
                        ev_list.sort(key=lambda x: x['start'])
                        
                        # 2. Fusion
                        merged_events = []
                        if ev_list:
                            curr = ev_list[0].copy() # Copy to avoid modifying original data
                            
                            for i in range(1, len(ev_list)):
                                nxt = ev_list[i]
                                # Si la fin de l'actuel == début du suivant (continuité parfaite)
                                if curr['end'] == nxt['start']:
                                    # On étend la fin
                                    curr['end'] = nxt['end']
                                    # On fusionne les groupes (union)
                                    g_curr = set(curr.get('groups', []))
                                    g_nxt = set(nxt.get('groups', []))
                                    curr['groups'] = sorted(list(g_curr | g_nxt))
                                else:
                                    merged_events.append(curr)
                                    curr = nxt.copy()
                            merged_events.append(curr)
                        
                        body += f"\nMatière : {subj}\n"
                        
                        for ev in merged_events:
                            dt = ev['start']
                            day_str = f"{days[dt.weekday()]} {dt.day} {months[dt.month]}"
                            h_start = dt.strftime("%Hh%M")
                            h_end = ev['end'].strftime("%Hh%M")
                            dur = (ev['end'] - ev['start']).total_seconds()/3600
                            total_h_global += dur
                            
                            grp_txt = ""
                            if ev['groups'] and len(ev['groups']) == 1:
                                grp_txt = f" ({ev['groups'][0]})"
                            
                            body += f"- {day_str} de {h_start} à {h_end}{grp_txt}\n"
                
                final_txt = f"{intro}\n{body}\nTotal planifié : {total_h_global:g} heures.\n\n{closing}"
                st.text_area("Aperçu (Texte brut pour mail)", value=final_txt, height=500)

    # --- TAB 3: RÉCAPITULATIFS ---
    with tab_stats:
        st.markdown('<div class="cesi-tag">Analyse</div>', unsafe_allow_html=True)
        st.write("")
        tabs_promo = st.tabs(list(events_map.keys()))
        
        for i, promo in enumerate(events_map.keys()):
            with tabs_promo[i]:
                evs_promo = events_map[promo]
                mode = st.radio(f"Vue {promo}", ["Par Matière", "Par Enseignant"], key=f"rad_{promo}", horizontal=True)
                
                if mode == "Par Matière":
                    subjs = sorted(list(set(e['summary'] for e in evs_promo)))
                    sel = st.selectbox(f"Matière ({promo})", subjs, key=f"sb_m_{promo}")
                    if sel:
                        data = []
                        for e in evs_promo:
                            if e['summary'] == sel:
                                data.append({
                                    "Date": e['start'].strftime("%d/%m/%Y"),
                                    "Heures": f"{e['start'].strftime('%H:%M')} - {e['end'].strftime('%H:%M')}",
                                    "Enseignant": ", ".join(e['teachers']),
                                    "Groupes": ", ".join(e['groups'])
                                })
                        st.dataframe(pd.DataFrame(data), use_container_width=True)
                
                else:
                    teachs = sorted(list(set(t for e in evs_promo for t in e['teachers'])))
                    sel = st.selectbox(f"Enseignant ({promo})", teachs, key=f"sb_t_{promo}")
                    if sel:
                        data = []
                        for e in evs_promo:
                            if sel in e['teachers']:
                                data.append({
                                    "Date": e['start'].strftime("%d/%m/%Y"),
                                    "Heures": f"{e['start'].strftime('%H:%M')} - {e['end'].strftime('%H:%M')}",
                                    "Matière": e['summary'],
                                    "Groupes": ", ".join(e['groups'])
                                })
                        st.dataframe(pd.DataFrame(data), use_container_width=True)

    # --- TAB 4: EXAMENS ---
    with tab_exam:
        st.markdown('<div class="cesi-tag">Examens</div>', unsafe_allow_html=True)
        st.write("")
        
        p1_exams = []
        p2_exams = []
        
        for e in all_events_flat:
            if e['description'] and "EXAMEN" in e['description'].upper():
                row = {
                    "Date": e['start'].strftime("%d/%m/%Y"),
                    "Horaire": f"{e['start'].strftime('%H:%M')} - {e['end'].strftime('%H:%M')}",
                    "Matière": e['summary'],
                    "Description": e['description']
                }
                if e['promo_label'] == "P1":
                    p1_exams.append(row)
                elif e['promo_label'] == "P2":
                    p2_exams.append(row)
        
        c_ex1, c_ex2 = st.columns(2)
        with c_ex1:
            st.markdown("#### Promo P1")
            if p1_exams:
                st.dataframe(pd.DataFrame(p1_exams), hide_index=True)
            else:
                st.info("Aucun examen détecté.")
        
        with c_ex2:
            st.markdown("#### Promo P2")
            if p2_exams:
                st.dataframe(pd.DataFrame(p2_exams), hide_index=True)
            else:
                st.info("Aucun examen détecté.")

    # --- TAB 5: EXPORTS ---
    with tab_export:
        st.markdown('<div class="cesi-tag">Exports</div>', unsafe_allow_html=True)
        st.write("")
        
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("#### Par Promo (Global)")
            for promo, evs in events_map.items():
                if evs:
                    # Export standard sans préfixe
                    ics = generate_ics(evs, include_prefix=False)
                    st.download_button(f"📥 {promo}.ics", data=ics, file_name=f"{promo}.ics", mime="text/calendar")
        with c2:
            st.markdown("#### Par Enseignant")
            if sorted_teachers:
                sel = st.multiselect("Choix", sorted_teachers, key="exp_sel")
                if st.button("Générer"):
                    # Filtrage des events
                    evs = [e for e in all_events_flat if any(t in e['teachers'] for t in sel)]
                    
                    # Génération AVEC PREFIXE [P1 G1]
                    ics = generate_ics(evs, include_prefix=True)
                    
                    # Nom du fichier basé sur le premier enseignant sélectionné
                    if len(sel) > 0:
                        # Nettoyage du nom (enlever les espaces) pour le fichier
                        safe_name = sel[0].replace(" ", "_").replace(",", "")
                        fname = f"Planning_{safe_name}.ics"
                    else:
                        fname = "Planning_Enseignants.ics"
                        
                    st.download_button(f"📥 Télécharger {fname}", data=ics, file_name=fname, mime="text/calendar")

    # --- TAB 6: MAQUETTE (LOGIQUE AVANCÉE) ---
    with tab_maquette:
        st.markdown('<div class="cesi-tag">Suivi Pédagogique</div>', unsafe_allow_html=True)
        
        if maquette_sheet:
            # 1. Lecture de la maquette brute pour avoir l'ordre et les cibles
            # Adaptation: on lit la maquette et on crée une structure propre comme dans l'ancienne appli
            raw_mq_df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=maquette_sheet, header=None, engine='openpyxl')
            
            # Reconstruction d'un DataFrame propre 'maquette_df' avec subject et target
            # Hypothèse colonne C (idx 2) = Matière, colonne M (idx 12) = Cible
            maquette_rows = []
            if raw_mq_df.shape[1] > 12:
                for i in range(len(raw_mq_df)):
                    subj = raw_mq_df.iat[i, 2]
                    tgt = raw_mq_df.iat[i, 12]
                    if pd.notna(subj) and str(subj).strip():
                        try: val = float(tgt)
                        except: val = None # None si pas de cible (header ou autre)
                        maquette_rows.append({'subject': str(subj).strip(), 'target': val})
            maquette_df = pd.DataFrame(maquette_rows)

            # 2. Sélecteurs Promo & Groupe
            c_mq1, c_mq2 = st.columns(2)
            with c_mq1:
                p_comp = st.selectbox("Comparer Promo (Feuille)", list(events_map.keys()), key="mq_promo")
            with c_mq2:
                g_comp = st.selectbox("Sélectionner le groupe", options=['G 1', 'G 2'], index=0, key="mq_group")

            # 3. Logique de filtrage (Legacy)
            def parse_group_sel(sel):
                s = sel.strip().upper().replace(' ', '')
                if s in ['G1','G 1']:
                    return {'G 1','G1'}
                if s in ['G2','G 2']:
                    return {'G 2','G2'}
                return {sel}

            sel_groups = parse_group_sel(g_comp)

            IGNORE_SUBJECTS = {
                "erasmus day",
                "forum international",
                "période entreprise",
                "férié",
                "mission à l'international",
                "matière",
                "matières",
                "divers"
            }

            # 4. Calcul des heures (Legacy Logic adapted to New Event Structure)
            def sum_hours_by_subject_and_group(events, groups_filter):
                totals = {}
                counts = {}
                for ev in events:
                    subj = ev['summary']
                    if subj and subj.strip().lower() in IGNORE_SUBJECTS:
                        continue 

                    # Filtrage par groupe
                    # Dans le parser, ev['groups'] est une liste/set de strings (ex: ['G 1', 'G 2'] ou ['G 1'])
                    ev_groups_norm = set([g.strip().upper().replace(' ', '') for g in ev.get('groups', [])])
                    
                    # Si l'événement n'a pas de groupe assigné (liste vide), on considère souvent que c'est promo entière
                    # Mais pour coller à la logique stricte demandée :
                    if not ev.get('groups'):
                        # Si vide -> Promo entière -> Compte pour G1 et G2
                        matches = True
                    else:
                        # Si groupes définis -> Doit intersecter avec la sélection
                        tgt_norm = {g.strip().upper().replace(' ', '') for g in groups_filter}
                        matches = len(ev_groups_norm.intersection(tgt_norm)) > 0
                    
                    if not matches:
                        continue

                    # Durée
                    delta = (ev['end'] - ev['start']).total_seconds() / 3600.0
                    totals[subj] = totals.get(subj, 0) + delta
                    counts[subj] = counts.get(subj, 0) + 1
                return totals, counts

            evs_target = events_map.get(p_comp, [])
            totals_by_subject, counts_by_subject = sum_hours_by_subject_and_group(evs_target, sel_groups)

            # 5. Construction du tableau final (basé sur l'ordre maquette)
            rows_out = []
            for _, row in maquette_df.iterrows():
                subj = row['subject']
                if subj.lower() in IGNORE_SUBJECTS:
                    continue

                target = row['target']
                hours = totals_by_subject.get(subj, 0.0)
                sessions = counts_by_subject.get(subj, 0)
                
                diff = None
                if target is not None and not pd.isna(target):
                    diff = hours - float(target)
                
                rows_out.append({
                    'subject': subj,
                    'target_hours': target,
                    'entered_hours': round(hours, 2),
                    'diff_hours': round(diff, 2) if diff is not None else None,
                    'sessions_entered': sessions
                })

            out_df = pd.DataFrame(rows_out, columns=['subject','target_hours','entered_hours','diff_hours','sessions_entered'])

            # 6. Affichage & Styling
            st.markdown(f"### Résultats pour **{p_comp}** — **{g_comp}**")
            st.caption("Ordre défini par le fichier Maquette. Les matières ignorées sont masquées.")

            def highlight_row(r):
                if r['target_hours'] is None or pd.isna(r['target_hours']) or r['target_hours'] == 0:
                    # Jaune si pas de cible (souvent titre ou optionnel)
                    if r['entered_hours'] > 0: return ['']*len(r) # Si ya des heures mais pas de cible, on laisse normal ou warning? Le code legacy mettait jaune si target is None
                    return ['background-color: #fff3cd']*len(r)
                
                # Rouge si écart significatif
                if r['diff_hours'] is not None and abs(r['diff_hours']) > 0.001:
                    return ['background-color: #f8d7da']*len(r)
                return ['']*len(r)

            st.dataframe(out_df.style.apply(highlight_row, axis=1), use_container_width=True)

            # 7. Résumé
            c_r1, c_r2, c_r3 = st.columns(3)
            total_expected = out_df['target_hours'].sum()
            total_entered = out_df['entered_hours'].sum()
            
            with c_r1:
                st.metric("Heures Maquette", f"{total_expected:.2f} h")
            with c_r2:
                st.metric("Heures Saisies", f"{total_entered:.2f} h")
            with c_r3:
                delta = total_entered - total_expected
                st.metric("Écart Global", f"{delta:+.2f} h", delta_color="inverse")

        else:
            st.warning("Pas de feuille Maquette détectée dans le fichier.")

    # --- TAB 7: VOLUME HORAIRE ---
    with tab_volume:
        st.markdown('<div class="cesi-tag">Volume Horaire</div>', unsafe_allow_html=True)
        st.write("")

        # --- Sélection de la période ---
        all_dates = [e['start'].date() for e in all_events_flat]
        min_date = min(all_dates) if all_dates else date.today()
        max_date = max(all_dates) if all_dates else date.today()

        col_d1, col_d2 = st.columns(2)
        with col_d1:
            date_debut = st.date_input("Date de début", value=min_date, min_value=min_date, max_value=max_date, key="vh_d1")
        with col_d2:
            date_fin = st.date_input("Date de fin", value=max_date, min_value=min_date, max_value=max_date, key="vh_d2")

        if date_debut > date_fin:
            st.error("La date de début doit être antérieure à la date de fin.")
            st.stop()

        # --- Exclusion de matières ---
        MATIERES_EXCLUES_DEFAUT = {
            "periode entreprise", "periode en entreprise",
            "periode alternance", "alternance",
            "ferie", "ferie", "ferie",
            "erasmus day", "forum international", "mission a l'international",
            "divers", "matiere", "matieres"
        }

        all_subjects_sorted = sorted(set(e['summary'] for e in all_events_flat))

        import unicodedata
        def normalize_str(s):
            s = s.strip().lower()
            s = unicodedata.normalize('NFD', s)
            s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
            return s

        default_excluded = [
            s for s in all_subjects_sorted
            if normalize_str(s) in MATIERES_EXCLUES_DEFAUT
        ]

        with st.expander("Matières exclues du calcul", expanded=True):
            st.caption("Ces matières sont exclues par défaut. Ajoutez-en ou retirez-en selon vos besoins.")
            matieres_exclues = st.multiselect(
                "Matières à exclure",
                options=all_subjects_sorted,
                default=default_excluded,
                key="vh_exclusions"
            )

        matieres_exclues_norm = {normalize_str(m) for m in matieres_exclues}

        # --- Filtrage sur la période + exclusions ---
        evs_periode = [
            e for e in all_events_flat
            if date_debut <= e['start'].date() <= date_fin
            and normalize_str(e['summary']) not in matieres_exclues_norm
        ]

        if not evs_periode:
            st.info("Aucune séance sur cette période.")
        else:
            # -----------------------------------------------------------------------
            # Fonctions de calcul
            # -----------------------------------------------------------------------

            def duree_h(ev):
                return (ev['end'] - ev['start']).total_seconds() / 3600.0

            def is_autonomie(ev):
                """Séance sans enseignant assigné → autonomie."""
                return len(ev.get('teachers', [])) == 0

            def nb_groupes_actifs(ev):
                """Nombre de demi-groupes distincts dans l'événement (min 1)."""
                return max(1, len(ev.get('groups', [])))

            def nb_formateurs(ev):
                """Nombre de formateurs dans l'événement."""
                return max(1, len(ev.get('teachers', [])))

            def heures_formation_event(ev):
                """
                Heures de formation = durée × nb_groupes.
                1h en dédoublement G1+G2 = 2h de formation (1h reçue par chaque groupe).
                """
                return duree_h(ev) * nb_groupes_actifs(ev)

            def heures_avec_enseignant_event(ev):
                """
                Heures avec enseignant = durée × nb_groupes, hors autonomie.
                1h cours classe entière = 1h avec enseignant.
                1h dédoublement G1+G2   = 2h avec enseignant (1h par groupe).
                Toujours >= heures_formateur car on compte les groupes, pas les formateurs.
                """
                if is_autonomie(ev):
                    return 0.0
                return duree_h(ev) * nb_groupes_actifs(ev)

            def heures_formateur_event(ev):
                """
                Heures formateur = durée × nb_formateurs.
                Reflète le temps réel passé par les formateurs, indépendamment des groupes.
                - 1h, 1 formateur, classe entière  → 1h formateur
                - 1h, 1 formateur, dédoublement    → 1h formateur  (même créneau, même formateur)
                - 1h, 2 formateurs simultanés       → 2h formateur
                """
                if is_autonomie(ev):
                    return 0.0
                return duree_h(ev) * nb_formateurs(ev)

            def heures_formation_group(evs, promo, groupe):
                """Heures de formation reçues par un groupe précis."""
                total = 0.0
                for ev in evs:
                    if ev.get('promo_label') != promo:
                        continue
                    groups = ev.get('groups', [])
                    if not groups:
                        total += duree_h(ev)
                    elif groupe in groups:
                        total += duree_h(ev)
                return total

            def heures_avec_enseignant_group(evs, promo, groupe):
                """Heures avec enseignant reçues par un groupe précis (hors autonomie)."""
                total = 0.0
                for ev in evs:
                    if ev.get('promo_label') != promo or is_autonomie(ev):
                        continue
                    groups = ev.get('groups', [])
                    if not groups:
                        total += duree_h(ev)
                    elif groupe in groups:
                        total += duree_h(ev)
                return total

            def heures_autonomie_group(evs, promo, groupe):
                """Heures en autonomie pour un groupe précis."""
                total = 0.0
                for ev in evs:
                    if ev.get('promo_label') != promo or not is_autonomie(ev):
                        continue
                    groups = ev.get('groups', [])
                    if not groups:
                        total += duree_h(ev)
                    elif groupe in groups:
                        total += duree_h(ev)
                return total

            # -----------------------------------------------------------------------
            # Calculs globaux
            # -----------------------------------------------------------------------

            total_formation       = sum(heures_formation_event(e) for e in evs_periode)
            total_avec_enseignant = sum(heures_avec_enseignant_event(e) for e in evs_periode)
            total_autonomie       = sum(heures_formation_event(e) for e in evs_periode if is_autonomie(e))
            total_formateur       = sum(heures_formateur_event(e) for e in evs_periode)

            # Par promo / groupe
            promos = sorted(events_map.keys())
            groupes_labels = ['G 1', 'G 2']

            # -----------------------------------------------------------------------
            # Affichage : Métriques globales
            # -----------------------------------------------------------------------
            st.markdown("### Vue d'ensemble de la période")
            st.caption(f"Du **{date_debut.strftime('%d/%m/%Y')}** au **{date_fin.strftime('%d/%m/%Y')}**")
            st.write("")

            cols_global = st.columns(4)
            metrics_global = [
                ("Heures de formation", total_formation, "Somme des heures × groupes"),
                ("Heures formateur", total_formateur, "Dédoublements & multi-formateurs pris en compte"),
                ("Heures en autonomie", total_autonomie, "Séances sans enseignant"),
                ("Heures avec enseignant", total_avec_enseignant, "Séances avec au moins 1 formateur"),
            ]
            for col, (label, val, help_txt) in zip(cols_global, metrics_global):
                with col:
                    st.markdown(f"""
                    <div class="metric-card" title="{help_txt}">
                        <div class="metric-value">{val:.1f} h</div>
                        <div class="metric-label">{label}</div>
                    </div>
                    """, unsafe_allow_html=True)

            st.write("")
            st.markdown("---")

            # -----------------------------------------------------------------------
            # Affichage : Détail par Promo & Groupe
            # -----------------------------------------------------------------------
            st.markdown("### Détail par Promo et Groupe")
            st.write("")

            detail_rows = []
            for promo in promos:
                evs_promo = [e for e in evs_periode if e.get('promo_label') == promo]
                if not evs_promo:
                    continue

                for grp in groupes_labels:
                    h_form    = heures_formation_group(evs_periode, promo, grp)
                    h_ens     = heures_avec_enseignant_group(evs_periode, promo, grp)
                    h_auto    = heures_autonomie_group(evs_periode, promo, grp)
                    # Heures formateur : durée × nb_formateurs, sur les séances de ce promo/groupe
                    h_formateur_grp = 0.0
                    for ev in evs_promo:
                        if is_autonomie(ev):
                            continue
                        groups = ev.get('groups', [])
                        if not groups or grp in groups:
                            h_formateur_grp += duree_h(ev) * nb_formateurs(ev)

                    detail_rows.append({
                        "Promo": promo,
                        "Groupe": grp,
                        "Heures formation": round(h_form, 2),
                        "Heures avec enseignant": round(h_ens, 2),
                        "Heures autonomie": round(h_auto, 2),
                        "Heures formateur": round(h_formateur_grp, 2),
                    })

            if detail_rows:
                df_detail = pd.DataFrame(detail_rows)

                # Affichage tableau stylé
                def style_detail(df):
                    return df.style.format({
                        "Heures formation": "{:.2f} h",
                        "Heures formateur": "{:.2f} h",
                        "Heures autonomie": "{:.2f} h",
                        "Heures avec enseignant": "{:.2f} h",
                    }).set_properties(**{
                        'text-align': 'center'
                    }).set_table_styles([
                        {'selector': 'th', 'props': [('text-align', 'center'), ('font-weight', 'bold')]}
                    ])

                st.dataframe(style_detail(df_detail), use_container_width=True, hide_index=True)

                # Totaux par promo (synthèse)
                st.write("")
                st.markdown("#### Synthèse par Promo (heures formateur, toutes promos)")
                synth_rows = []
                for promo in promos:
                    evs_promo_all = [e for e in evs_periode if e.get('promo_label') == promo]
                    if not evs_promo_all:
                        continue
                    synth_rows.append({
                        "Promo": promo,
                        "Heures formation (Σ groupes)": round(
                            sum(heures_formation_event(e) for e in evs_promo_all), 2),
                        "Heures avec enseignant (Σ groupes)": round(
                            sum(heures_avec_enseignant_event(e) for e in evs_promo_all), 2),
                        "Heures autonomie (Σ groupes)": round(
                            sum(heures_formation_event(e) for e in evs_promo_all if is_autonomie(e)), 2),
                        "Heures formateur": round(
                            sum(heures_formateur_event(e) for e in evs_promo_all), 2),
                    })
                if synth_rows:
                    df_synth = pd.DataFrame(synth_rows)
                    st.dataframe(df_synth.style.format({
                        col: "{:.2f} h" for col in df_synth.columns if "Heure" in col
                    }), use_container_width=True, hide_index=True)

            # -----------------------------------------------------------------------
            # Aide / Légende
            # -----------------------------------------------------------------------
            st.write("")
            with st.expander("ℹ️ Explication des calculs"):
                st.markdown("""
**Heures de formation** : total des heures planifiées, chaque groupe étant compté séparément.  
*Exemple : 1h en dédoublement G1/G2 = 2h de formation (1h reçue par chaque groupe).*

**Heures avec enseignant** : même logique que les heures de formation, mais uniquement pour les séances avec au moins un formateur.  
*Toujours ≤ heures de formation (la différence = autonomie).*

**Heures en autonomie** : séances sans enseignant assigné (travail personnel, e-learning…).  
*Heures de formation = heures avec enseignant + heures en autonomie.*

**Heures formateur** : temps réellement passé par les formateurs, indépendamment des groupes.  
- 1h, 1 formateur, classe entière → **1h formateur**  
- 1h, 1 formateur, dédoublement G1+G2 → **1h formateur** (même créneau, même formateur)  
- 1h, 2 formateurs simultanés → **2h formateur** (chaque formateur compte)  

*Invariant garanti : heures formateur ≤ heures avec enseignant.*
                """)


else:
    # État initial (sans fichier)
    st.markdown("""
    <div style="text-align: center; margin-top: 50px; padding: 40px; background-color: #FFFFFF; border-radius: 12px; border: 1px solid #E5E5E5;">
        <div class="cesi-title">Modern Opus</div>
        <p style="color: #666666; font-size: 1.1em; margin-top:20px;">
            Déposez votre fichier Excel pour générer les plannings, mails et analyses.
        </p>
        <div style="margin-top: 30px;">
            <span class="cesi-tag">P1</span>
            <span class="cesi-tag">P2</span>
            <span class="cesi-tag">Maquette</span>
        </div>
    </div>
    """, unsafe_allow_html=True)
